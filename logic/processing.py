import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import google.generativeai as genai
import requests
import undetected_chromedriver as uc
from g4f.client import Client
from markdownify import markdownify as md
from selenium.common.exceptions import WebDriverException

from logic.models import ProcessingContext, Task

DATA_DIR = "data/results"

SELECTORS_TO_REMOVE = (
    "script, style, noscript, iframe, header, footer, nav, aside,"
    ".header, .footer, .nav, .menu, .sidebar, .ads, .advertisement,"
    ".social, .breadcrumbs, .comments, .related, .popup, .subscribe,"
    ".newsletter, .cookie, .cookie-banner, .modal, #comments, #footer, #header"
    "img, picture, a"
)

# Instantiate the client once at the module level for reuse and performance.
g4f_client = Client()

# Configure the Gemini client.
# It's good practice to configure the API key once at the module level.
# The user should set the GOOGLE_API_KEY environment variable.
if "GOOGLE_API_KEY" in os.environ:
    genai.configure(api_key=os.environ["GOOGLE_API_KEY"])
else:
    print("Warning: GOOGLE_API_KEY environment variable not set. Gemini models will not work.")


def save_to_excel(
    task: Task,
    md_content,
    processed_content,
    run_id: str,
):
    """Saves the provided data to a daily Excel file in the data/results directory."""
    try:
        os.makedirs(DATA_DIR, exist_ok=True)
        excel_file = os.path.join(DATA_DIR, f"results_{run_id}.xlsx")

        write_header = not os.path.exists(excel_file)

        if write_header:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Processed Data"
            header = [
                "Domain", "Source Estate ID", "Source ID", "URL",
                "Status", "Rent Status", "Subtype", "Type", "Processed Content",
            ]
            sheet.append(header)
        else:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

        # Append the new row of data
        sheet.append([
            task.domain, task.source_estate_id, task.source_id, task.url,
            task.status, task.rent_status, task.subtype, task.type,
            processed_content,
        ])

        # Set text wrapping for the "Processed Content" cell in the newly added row
        last_row_num = sheet.max_row
        processed_content_cell = sheet.cell(row=last_row_num, column=9)  # Column 'I'
        processed_content_cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Adjust column widths for readability
        for i, column_cells in enumerate(sheet.columns, 1):
            column_letter = get_column_letter(i)

            if column_letter == 'I':  # Processed Content
                sheet.column_dimensions[column_letter].width = 80
            elif column_letter == 'D':  # URL
                sheet.column_dimensions[column_letter].width = 60
            else:
                # Calculate max length for other columns
                max_length = max(
                    (len(str(cell.value)) for cell in column_cells if cell.value),
                    default=0
                )
                # Set a slightly larger width, with a reasonable minimum for headers
                adjusted_width = max(max_length + 2, 15)
                sheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(excel_file)
        return True, f"Saved to {os.path.basename(excel_file)}"
    except (IOError, openpyxl.utils.exceptions.InvalidFileException) as e:
        return False, f"Failed to save to Excel: {e}"
    except Exception as e:
        return False, f"An unexpected error occurred while saving to Excel: {e}"


def _process_and_save_markdown(
    md_content: str,
    task: Task,
    context: ProcessingContext,
    success_message_prefix: str,
):
    """Helper to process MD, update UI, and save results."""
    context.ui_queue.put(("update_text", ("raw", md_content)))

    processed_text = process_md(
        md_content,
        context.user_prompt_template,
        context.system_prompt_text,
        context.model_name,
    )
    context.ui_queue.put(("update_text", ("processed", processed_text)))

    status_message = success_message_prefix
    if context.save_excel:
        success, message = save_to_excel(
            task, md_content, processed_text, context.run_id
        )
        status_message += f" | {message}"
        if not success:
            context.ui_queue.put(("error", message))

    context.ui_queue.put(("update_status", status_message))


def fetch_md(task: Task, context: ProcessingContext):
    """Fetches markdown content using the Jina Reader API."""
    if not task.url:
        context.ui_queue.put(("error", "Encountered a task with no URL."))
        return

    try:
        headers = {
            "Authorization": f"Bearer {context.api_key}",
            "Content-Type": "application/json",
            "X-Exclude-Selector": ",".join(SELECTORS_TO_REMOVE),
        }

        if context.use_proxy and context.proxy_url:
            headers["X-Proxy-Url"] = context.proxy_url

        response = requests.get(
            f"https://r.jina.ai/{task.url}", headers=headers, timeout=30
        )

        if response.status_code == 200:
            md_content = response.text
            _process_and_save_markdown(
                md_content, task, context, "Completed successfully"
            )
        else:
            error_msg = f"API Error {response.status_code}: {response.text}"
            context.ui_queue.put(("error", error_msg))
            context.ui_queue.put(("update_text", ("raw", error_msg)))

    except requests.exceptions.RequestException as e:
        error_msg = f"Request failed: {str(e)}"
        context.ui_queue.put(("error", error_msg))
        context.ui_queue.put(("update_text", ("raw", error_msg)))


def fetch_md_selenium(task: Task, context: ProcessingContext):
    """Fetches HTML content using Selenium and converts it to markdown."""
    if not task.url:
        context.ui_queue.put(("error", "Encountered a task with no URL."))
        return

    driver = None
    try:
        chrome_options = uc.ChromeOptions()
        # chrome_options.add_argument('''--headless''')
        chrome_options.add_argument("--disable-gpu")

        if context.use_proxy and context.proxy_url:
            chrome_options.add_argument(f"--proxy-server={context.proxy_url}")

        driver = uc.Chrome(options=chrome_options, use_subprocess=True)
        driver.get(task.url)
        # A simple wait for elements to appear.
        # For more complex pages, explicit waits (WebDriverWait) are more robust.
        driver.implicitly_wait(5)

        # JavaScript to remove elements matching the selectors.
        # This helps in cleaning the HTML before converting to Markdown.
        js_remover_script = """
        const selectors = arguments[0].split(',');
        for (const selector of selectors) {
            try {
                document.querySelectorAll(selector.trim()).forEach(el => el.remove());
            } catch (e) {
                // Silently ignore errors for invalid selectors
            }
        }
        """
        driver.execute_script(js_remover_script, ",".join(SELECTORS_TO_REMOVE))

        html_content = driver.page_source

        md_content = md(
            html_content, strip=["a", "img", "script", "style", "svg", "button"]
        )

        # Clean up excessive whitespace and blank lines from the markdown content.
        lines = [line.strip() for line in md_content.splitlines()]
        cleaned_md_content = "\n".join(line for line in lines if line)

        _process_and_save_markdown(
            cleaned_md_content,
            task,
            context,
            "Completed successfully via Selenium",
        )

    except (WebDriverException, Exception) as e:
        error_msg = f"Selenium failed: {str(e)}"
        context.ui_queue.put(("error", error_msg))
        context.ui_queue.put(("update_text", ("raw", error_msg)))
    finally:
        if driver:
            driver.quit()


def process_md(raw_md, user_prompt_template, system_prompt_text, model_name: str):
    user_prompt = user_prompt_template.format(content=raw_md)
    system_prompt = system_prompt_text.strip()

    # Logic for Gemini models using the official Google library
    if model_name.startswith("gemini"):
        try:
            if not os.environ.get("GOOGLE_API_KEY"):
                return "Error: GOOGLE_API_KEY environment variable not set. Please configure it to use Gemini."

            model = genai.GenerativeModel(
                model_name=model_name, system_instruction=system_prompt
            )
            response = model.generate_content(user_prompt)
            return response.text
        except Exception as e:
            return f"An error occurred with the Gemini API: {e}"

    # Existing logic for g4f models (GPT, Claude, etc.)
    else:
        try:
            messages = [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ]
            response = g4f_client.chat.completions.create(
                model=model_name, messages=messages, web_search=False
            )
            return response.choices[0].message.content
        except Exception as e:
            return f"An error occurred with the g4f client: {e}"
